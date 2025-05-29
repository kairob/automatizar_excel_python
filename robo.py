from openpyxl import Workbook, load_workbook


grade= load_workbook('grade.xlsx')
grade_003  = grade['grade003']

for row in grade_003.iter_rows(values_only=True):
    print(row)